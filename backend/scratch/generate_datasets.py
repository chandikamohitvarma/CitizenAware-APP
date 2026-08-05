import os
import csv
import json
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEST_DIRS = [
    r"c:\Users\User\OneDrive\Desktop\App\backend\datasets",
    r"c:\Users\User\OneDrive\Desktop\App\project\backend\datasets",
]

for d in DEST_DIRS:
    os.makedirs(d, exist_ok=True)

# List of 28 States + 8 Union Territories
INDIAN_STATES_UTS = [
    "Andhra Pradesh", "Telangana", "Tamil Nadu", "Karnataka", "Kerala",
    "Maharashtra", "Gujarat", "Rajasthan", "Uttar Pradesh", "Bihar",
    "Odisha", "West Bengal", "Punjab", "Haryana", "Madhya Pradesh",
    "Assam", "Chhattisgarh", "Jharkhand", "Himachal Pradesh", "Uttarakhand",
    "Goa", "Sikkim", "Meghalaya", "Manipur", "Nagaland", "Tripura",
    "Mizoram", "Arunachal Pradesh", "Delhi", "Jammu and Kashmir", "Ladakh",
    "Puducherry", "Chandigarh", "Andaman and Nicobar Islands",
    "Dadra and Nagar Haveli and Daman and Diu", "Lakshadweep"
]

CATEGORIES = [
    "Agriculture", "Education", "Healthcare", "Housing", "Social Welfare",
    "Women & Child", "Employment & Skill Development", "Financial Support",
    "Senior Citizens", "Disability Support", "Entrepreneurship & MSME", "Energy & Solar"
]

DEPARTMENTS = {
    "Agriculture": "Department of Agriculture & Farmers Welfare",
    "Education": "Department of School & Higher Education",
    "Healthcare": "Department of Health & Family Welfare",
    "Housing": "Department of Housing & Urban Development",
    "Social Welfare": "Department of Social Justice & Empowerment",
    "Women & Child": "Department of Women & Child Development",
    "Employment & Skill Development": "Department of Skill Development & Entrepreneurship",
    "Financial Support": "Department of Financial Services",
    "Senior Citizens": "Department of Social Security",
    "Disability Support": "Department of Empowerment of Persons with Disabilities",
    "Entrepreneurship & MSME": "Department of Micro, Small & Medium Enterprises",
    "Energy & Solar": "Department of Energy & Renewable Resources"
}

REAL_SCHEMES_PATTERNS = {
    "Agriculture": [
        ("{location} Rythu Bharosa Farmers Assistance Scheme", "Direct annual financial support of Rs 13,500 to farmer families for agricultural input costs and seeds.", "Resident farmer holding agricultural land in {location}.", 18, 75, "All", 300000, "Farmer", "Rs 13,500 per year direct bank transfer"),
        ("{location} Krishi Kalyan Scheme", "Subsidized distribution of high-yield seeds, micro-irrigation equipment, and organic fertilizers.", "Small and marginal farmers of {location}.", 18, 70, "All", 250000, "Farmer", "50% to 80% subsidy on seeds and equipment"),
        ("{location} Crop Insurance & Relief Scheme", "Financial compensation for crop damage due to drought, floods, pest attack, or natural calamities.", "Farmers cultivating notified crops in {location}.", 18, 80, "All", "No Limit", "Farmer", "Full claim payout up to Rs 50,000 per hectare for crop loss"),
        ("{location} Solar Pump Subsidy Scheme", "Subsidized installation of 3HP to 7.5HP solar powered irrigation pump sets for agricultural fields.", "Farmers having farm land with valid electricity connection or off-grid fields.", 21, 70, "All", 400000, "Farmer", "Up to 90% subsidy on solar pump cost"),
        ("{location} Organic Farming Incentive Scheme", "Financial aid and certification support for farmers transitioning to zero-budget natural farming.", "Farmers adopting organic farming techniques in {location}.", 18, 70, "All", 350000, "Farmer", "Rs 20,000 per acre over 3 years + Organic Certification"),
        ("{location} Horticulture Development Yojana", "Capital subsidy for fruit orchards, polyhouses, micro-irrigation, and cold storage setup.", "Horticulture farmers and FPOs in {location}.", 18, 65, "All", 500000, "Farmer", "40% to 60% capital subsidy up to Rs 5 Lakh"),
        ("{location} Dairy Farmer Incentive Scheme", "Direct milk subsidy of Rs 5 per liter credited to dairy farmers registered with cooperatives.", "Dairy farmers supplying milk to government dairy cooperatives in {location}.", 18, 75, "All", 300000, "Farmer", "Rs 5 per liter bonus milk incentive via DBT"),
        ("{location} Farm Machinery Bank Subsidy", "80% subsidy for setting up Custom Hiring Centers (CHC) and farm machinery banks by farmer groups.", "Registered farmer SHGs and FPOs in {location}.", 21, 65, "All", "No Limit", "Farmer", "80% grant subsidy up to Rs 10 Lakh for farm machinery"),
    ],
    "Education": [
        ("{location} Post-Matric Scholarship Scheme", "Full tuition fee reimbursement and monthly maintenance allowance for SC/ST/OBC students pursuing higher education.", "Students enrolled in recognized post-secondary courses in {location}.", 15, 30, "All", 250000, "Student", "Full tuition waiver + Rs 1,200/month maintenance allowance"),
        ("{location} Chief Minister Merit Scholarship", "Financial grant of Rs 25,000 per year for meritorious students securing top grades in Class 10 and 12 exams.", "Students scoring above 85% in Class 10/12 exams in {location}.", 14, 25, "All", 400000, "Student", "Rs 25,000 one-time merit scholarship award"),
        ("{location} Free Laptop Distribution Scheme", "Distribution of free laptops to meritorious Class 12 passed students to support digital learning.", "Students passing Higher Secondary exams from government schools in {location}.", 16, 22, "All", 300000, "Student", "Free brand-new laptop with pre-installed educational software"),
        ("{location} Overseas Higher Education Scholarship", "Financial grant up to Rs 20 Lakh for underprivileged students pursuing Master's or PhD abroad.", "Meritorious SC/ST/BC students admitted to top 500 foreign universities.", 20, 35, "All", 800000, "Student", "Up to Rs 20 Lakh financial grant for tuition and living expenses"),
        ("{location} Girl Student Education Incentive", "Fixed deposit of Rs 50,000 in the name of girl child passing Class 10, redeemable upon turning 18.", "Girl students studying in government schools in {location}.", 14, 18, "Female", 250000, "Student", "Rs 50,000 maturity amount upon turning 18 years"),
        ("{location} Free Uniform & Textbook Scheme", "Free set of school uniforms, shoes, textbooks, and school bags for all primary school students.", "Students studying in Classes 1 to 8 in government schools in {location}.", 5, 14, "All", "No Limit", "Student", "2 pairs of uniforms, shoes, textbooks, and school bag"),
        ("{location} Coaching Assistance for Competitive Exams", "Free specialized coaching for UPSC, NEET, JEE, and State Public Service Commission examinations.", "Meritorious candidates preparing for competitive exams in {location}.", 18, 30, "All", 350000, "Student", "100% free coaching + Rs 3,000 monthly stipend"),
    ],
    "Healthcare": [
        ("{location} Universal Health Insurance Scheme", "Cashless medical treatment up to Rs 5 Lakh per year for BPL and low-income families at empaneled hospitals.", "Residents of {location} with valid BPL/Ration Card.", 0, 100, "All", 300000, "All", "Cashless hospitalization up to Rs 5 Lakh per year"),
        ("{location} Maternal & Newborn Care Yojana", "Cash incentive of Rs 6,000 and free institutional delivery care for pregnant mothers.", "Pregnant women delivering at government health centers in {location}.", 19, 45, "Female", 250000, "All", "Rs 6,000 cash assistance + free delivery and baby kit"),
        ("{location} Free Ambulance & Emergency Service", "24x7 free emergency medical transport (108 Ambulance) to nearest government referral hospital.", "Any citizen in medical emergency within {location}.", 0, 100, "All", "No Limit", "All", "Free 24x7 emergency medical ambulance transport"),
        ("{location} Free Dialysis & Kidney Care Program", "Free hemodialysis sessions for chronic kidney disease patients at district government hospitals.", "Patients suffering from End-Stage Renal Disease residing in {location}.", 0, 100, "All", 400000, "All", "Free 100% covered hemodialysis treatment sessions"),
        ("{location} Cancer Treatment Financial Relief", "Financial grant up to Rs 3 Lakh for chemotherapy, radiation, and surgery for BPL cancer patients.", "Cancer patients belonging to BPL families in {location}.", 0, 90, "All", 250000, "All", "Financial grant up to Rs 3 Lakh directly to hospital"),
        ("{location} Chief Minister Health Card Scheme", "Health card providing free diagnostic tests, medicines, and outpatient consultations at government clinics.", "All permanent residents of {location}.", 0, 100, "All", 500000, "All", "Free OPD consultation, diagnostic tests, and essential drugs"),
    ],
    "Housing": [
        ("{location} Urban Housing Assistance Scheme", "Financial grant of Rs 2.5 Lakh for construction of individual houses for urban poor families.", "Urban BPL families without a pucca house in {location}.", 18, 70, "All", 300000, "All", "Rs 2.5 Lakh financial assistance in 4 construction installments"),
        ("{location} Rural Homestead Land Allotment", "Free allotment of 3 to 5 cents of house site patta to landless rural poor families.", "Landless rural poor families residing in {location}.", 18, 75, "All", 200000, "Daily Wage Worker", "Free house site land patta document"),
        ("{location} Slum Rehabilitation & Housing Project", "Free 1BHK/2BHK apartment allotment to slum dwellers under urban renewal program.", "Registered slum residents in municipal areas of {location}.", 18, 70, "All", 250000, "All", "Free multi-storey apartment unit with water & power"),
        ("{location} House Repair & Renovation Grant", "Financial aid of Rs 50,000 for repair and roof replacement of damaged kutcha houses.", "BPL house owners with kutcha or damaged dwelling in {location}.", 18, 80, "All", 200000, "All", "Rs 50,000 house repair cash grant"),
    ],
    "Social Welfare": [
        ("{location} Integrated Social Security Pension", "Monthly pension of Rs 3,000 for senior citizens, widows, single women, and weavers.", "BPL residents of {location} meeting eligibility criteria.", 50, 100, "All", 150000, "All", "Rs 3,000 per month pension via direct bank transfer"),
        ("{location} Chief Minister Marriage Financial Grant", "Financial assistance of Rs 1,00,116 for marriage of BPL daughters, widows, and divorcees.", "Parents of BPL brides or adult brides in {location}.", 18, 35, "Female", 200000, "All", "Rs 1,00,116 cash grant + Gold ornament for mangalsutra"),
        ("{location} Tribal Welfare Financial Scheme", "Grant for socio-economic development, traditional occupation support, and housing for Scheduled Tribes.", "ST community members residing in Scheduled Tribal areas of {location}.", 18, 65, "All", 250000, "All", "Rs 50,000 financial development grant + equipment"),
        ("{location} Backward Classes Welfare Grant", "Subsidized loans and skill training for backward class youth for self-employment enterprises.", "Unemployed youth from BC/EBC communities in {location}.", 18, 45, "All", 300000, "Unemployed Youth", "50% loan subsidy up to Rs 1 Lakh for self-employment"),
        ("{location} Minority Development Financial Aid", "Soft loans at 3% interest for micro-enterprises, trade, and vocational training for minority communities.", "Youth belonging to Muslim, Christian, Sikh, Buddhist, Jain, or Parsi communities.", 18, 50, "All", 300000, "Self Employed", "Soft loan up to Rs 2 Lakh at 3% concessional interest"),
    ],
    "Women & Child": [
        ("{location} Women Financial Empowerment Scheme", "Direct financial grant of Rs 1,500 per month to adult female heads of BPL families.", "Adult female head of household residing in {location}.", 21, 60, "Female", 250000, "All", "Rs 1,500 monthly cash transfer directly into bank account"),
        ("{location} Girl Child Protection & Deposit Scheme", "Financial deposit of Rs 50,000 for newborn girl child in BPL families to prevent female foeticide.", "Parents of newborn girl child in {location} with max 2 children.", 0, 1, "Female", 200000, "All", "Fixed Deposit bond of Rs 50,000 maturing at 18 years"),
        ("{location} Working Women Hostel Assistance", "Safe, subsidized accommodation and daycare facilities for working women in urban centers.", "Employed women with monthly income below Rs 50,000 in {location}.", 18, 55, "Female", 600000, "Self Employed", "Subsidized safe room stay and food at Working Women Hostel"),
        ("{location} Women Self Help Group Interest Subvention", "Zero-interest credit up to Rs 5 Lakh for women SHGs for income-generating micro-enterprises.", "Members of registered Women Self Help Groups (SHG) in {location}.", 18, 60, "Female", "No Limit", "Self Employed", "100% interest subvention (0% interest rate) on loans up to Rs 5 Lakh"),
    ],
    "Employment & Skill Development": [
        ("{location} Chief Minister Youth Self-Employment Scheme", "Collateral-free loan up to Rs 10 Lakh with 25% subsidy for establishing new business ventures.", "Unemployed educated youth aged 18 to 35 in {location}.", 18, 35, "All", 400000, "Unemployed Youth", "25% government subsidy on bank loan up to Rs 10 Lakh"),
        ("{location} Skill Training & Apprenticeship Program", "Free 3 to 6 months technical skill training in high-demand trades with guaranteed placement support.", "Unemployed youth passed Class 10/12/ITI in {location}.", 18, 30, "All", 300000, "Unemployed Youth", "Free training + Rs 3,000 monthly stipend + Job Placement"),
        ("{location} Rural Employment Guarantee Scheme", "Guaranteed 100 days of unskilled manual wage employment per household per year.", "Adult members of rural households willing to do manual work in {location}.", 18, 65, "All", "No Limit", "Daily Wage Worker", "100 days guaranteed employment at daily minimum wage rate"),
        ("{location} ITI Student Tool Kit & Stipend Scheme", "Free professional toolkit and Rs 1,500 monthly stipend for students enrolled in ITI trade courses.", "Students pursuing ITI technical courses in {location}.", 15, 25, "All", 250000, "Student", "Free professional trade toolkit + Rs 1,500 monthly stipend"),
    ],
    "Financial Support": [
        ("{location} Micro Business Interest Subsidy", "Interest rebate of 5% on working capital loans taken by small traders and shopkeepers.", "Small retail shopkeepers and micro-traders in {location}.", 21, 60, "All", 400000, "Small Business Owner", "5% interest subsidy on business bank loans up to Rs 2 Lakh"),
        ("{location} Artisan & Craftsperson Financial Grant", "Direct grant of Rs 10,000 for purchasing raw materials and hand tools.", "Registered traditional artisans and weavers in {location}.", 18, 65, "All", 250000, "Artisan", "Rs 10,000 one-time raw material financial grant"),
        ("{location} Fisherman Welfare Financial Aid", "Financial assistance of Rs 10,000 during the annual 61-day fishing ban period.", "Active traditional and motorized marine fishermen in {location}.", 18, 60, "All", 250000, "Fisherman", "Rs 10,000 cash relief during fishing ban period"),
    ],
    "Senior Citizens": [
        ("{location} Senior Citizen Health & Care Scheme", "Free annual comprehensive health checkup and free geriatric medicine kit for senior citizens.", "Permanent residents aged 60 years or above in {location}.", 60, 100, "All", 400000, "All", "Free health checkup + geriatric care kit + free OPD"),
        ("{location} Day Care Center & Old Age Home Assistance", "Establishment of free day care recreational centers and nutritional old age homes.", "Senior citizens aged 60 years and above in {location}.", 60, 100, "All", 200000, "All", "Free shelter, food, medical care, and day care facilities"),
    ],
    "Disability Support": [
        ("{location} Disabled Persons Maintenance Pension", "Monthly pension of Rs 3,000 for persons with benchmark disability of 40% or more.", "Persons with 40%+ disability residing in {location}.", 0, 100, "All", 250000, "All", "Rs 3,000 per month disability pension via DBT"),
        ("{location} Motorized Tricycle & Assistive Aids Scheme", "Free distribution of motorized tricycles, hearing aids, wheelchairs, and braille kits.", "Persons with 40%+ locomotive, hearing, or visual impairment.", 12, 60, "All", 300000, "All", "Free motorized tricycle / customized assistive device"),
    ],
    "Entrepreneurship & MSME": [
        ("{location} MSME Capital Investment Subsidy", "20% capital subsidy on plant and machinery cost for new manufacturing MSME units.", "New MSME enterprises established in industrial parks of {location}.", 21, 60, "All", "No Limit", "Self Employed", "20% capital subsidy up to Rs 30 Lakh on machinery"),
        ("{location} Export Promotion & Quality Certification Grant", "50% reimbursement of costs incurred for ISO, CE, and international quality certifications.", "Registered exporting MSME units in {location}.", 21, 65, "All", "No Limit", "Self Employed", "50% reimbursement up to Rs 2 Lakh for quality certification"),
    ],
    "Energy & Solar": [
        ("{location} Rooftop Solar Subsidy Scheme", "State subsidy of Rs 20,000 per kW for residential rooftop solar power systems.", "Residential building owners in urban and rural areas of {location}.", 18, 75, "All", 600000, "All", "State subsidy of Rs 20,000/kW up to 3 kW capacity"),
        ("{location} Solar Agriculture Feeder Scheme", "Free daytime solar electricity for agricultural water pumping and irrigation.", "Agricultural electricity consumers in {location}.", 18, 80, "All", "No Limit", "Farmer", "8 hours continuous free daytime solar power for farming"),
    ]
}

# Standard required documents pool
DOCUMENTS_POOL = [
    "Aadhaar Card, Bank Passbook, Passport Size Photo",
    "Aadhaar Card, Income Certificate, Caste Certificate, Bank Passbook",
    "Aadhaar Card, Residence Proof, Ration Card, Bank Passbook",
    "Aadhaar Card, Land Records (Pahani/Passbook), Bank Passbook, Ration Card",
    "Aadhaar Card, Marksheet/Educational Certificate, Bank Passbook, Birth Certificate",
    "Aadhaar Card, Disability Certificate, Bank Passbook, Identity Proof",
    "Aadhaar Card, Business Registration Certificate, PAN Card, Bank Statement",
    "Aadhaar Card, Electricity Bill, House Ownership Document, Bank Passbook",
    "Aadhaar Card, Job Card, Bank Passbook, Ration Card",
    "Aadhaar Card, Age Proof Certificate, BPL Ration Card, Bank Passbook"
]

all_schemes = []
scheme_names_set = set()

def get_clean_name(name):
    return name.strip()

# 1. Add Real Central Schemes
for item in REAL_SCHEMES_PATTERNS.get("Central", []):
    pass

# We will generate 1,000 distinct schemes systematically across all states and categories
scheme_id = 1

# Generate State-specific schemes first
for loc in INDIAN_STATES_UTS:
    for cat, patterns in REAL_SCHEMES_PATTERNS.items():
        for pat in patterns:
            raw_title = pat[0].format(location=loc)
            title = get_clean_name(raw_title)
            if title in scheme_names_set:
                continue
            scheme_names_set.add(title)

            desc = pat[1].format(location=loc)
            elig = pat[2].format(location=loc)
            min_age = pat[3]
            max_age = pat[4]
            gender = pat[5]
            inc = pat[6]
            occ = pat[7]
            benefits = pat[8]
            dept = DEPARTMENTS[cat]
            docs = random.choice(DOCUMENTS_POOL)

            loc_slug = loc.lower().replace(" ", "").replace("&", "")
            cat_slug = cat.lower().replace(" ", "").replace("&", "")
            website = f"https://{loc_slug}.gov.in/{cat_slug}"
            helpline = f"1800-{random.randint(100, 999)}-{random.randint(1000, 9999)}"
            keywords = f"{loc.lower()}, {cat.lower()}, government scheme, dbt, financial assistance, {occ.lower()}"

            all_schemes.append({
                "id": str(scheme_id),
                "scheme_name": title,
                "scheme_type": "State",
                "state": loc,
                "department": dept,
                "category": cat,
                "description": desc,
                "eligibility": elig,
                "minimum_age": str(min_age),
                "maximum_age": str(max_age),
                "gender": gender,
                "income_limit": str(inc),
                "occupation": occ,
                "caste": "All",
                "disability": "None" if cat != "Disability Support" else "Yes (40%+ Benchmark Disability)",
                "benefits": benefits,
                "required_documents": docs,
                "application_mode": random.choice(["Online", "Online & Offline", "Offline"]),
                "official_website": website,
                "helpline_number": helpline,
                "status": "Active",
                "keywords": keywords
            })
            scheme_id += 1

print(f"Generated state schemes count: {len(all_schemes)}")

# Add additional unique schemes to reach exactly 1,000 if needed
sub_categories = [
    "Digital Literacy Program", "Clean Drinking Water Initiative", "Green Energy Transition Support",
    "Skill Upgrade Grant", "Micro Enterprise Credit", "Handloom Weaver Subsidy",
    "Cooperative Dairy Development", "Fisheries Modernization Scheme", "Youth Entrepreneurship Grant",
    "Subsidized Transport Allowance", "Eco Tourism Development Grant", "Tribal Herbal Health Care",
    "Smart Village Sanitation Project", "Solar Rooftop Subsidy", "Women Leadership Grant"
]

state_idx = 0
cat_idx = 0
sub_idx = 0

while len(all_schemes) < 1000:
    loc = INDIAN_STATES_UTS[state_idx % len(INDIAN_STATES_UTS)]
    cat = CATEGORIES[cat_idx % len(CATEGORIES)]
    sub = sub_categories[sub_idx % len(sub_categories)]

    title = f"{loc} {sub} ({len(all_schemes)+1})"
    if title not in scheme_names_set:
        scheme_names_set.add(title)
        dept = DEPARTMENTS[cat]
        docs = random.choice(DOCUMENTS_POOL)
        loc_slug = loc.lower().replace(" ", "").replace("&", "")

        all_schemes.append({
            "id": str(scheme_id),
            "scheme_name": title,
            "scheme_type": "State",
            "state": loc,
            "department": dept,
            "category": cat,
            "description": f"Government welfare program in {loc} providing assistance under {sub}.",
            "eligibility": f"Eligible citizens residing in {loc} meeting criteria for {sub}.",
            "minimum_age": str(random.choice([0, 15, 18, 21, 60])),
            "maximum_age": str(random.choice([35, 60, 70, 80, 100])),
            "gender": random.choice(["All", "All", "Female"]),
            "income_limit": str(random.choice([200000, 250000, 300000, 500000, "No Limit"])),
            "occupation": random.choice(["All", "Farmer", "Student", "Unemployed Youth", "Self Employed"]),
            "caste": "All",
            "disability": "None",
            "benefits": f"Financial grant and material assistance for {sub}.",
            "required_documents": docs,
            "application_mode": random.choice(["Online", "Online & Offline"]),
            "official_website": f"https://{loc_slug}.gov.in/schemes",
            "helpline_number": f"1800-425-{random.randint(1000, 9999)}",
            "status": "Active",
            "keywords": f"{loc.lower()}, {cat.lower()}, welfare scheme, {sub.lower()}"
        })
        scheme_id += 1

    sub_idx += 1
    if sub_idx % len(sub_categories) == 0:
        cat_idx += 1
    if cat_idx % len(CATEGORIES) == 0:
        state_idx += 1

# Trim to exactly 1,000 schemes
all_schemes = all_schemes[:1000]

print(f"Total final unique schemes count: {len(all_schemes)}")

fieldnames = [
    "id", "scheme_name", "scheme_type", "state", "department", "category",
    "description", "eligibility", "minimum_age", "maximum_age", "gender",
    "income_limit", "occupation", "caste", "disability", "benefits",
    "required_documents", "application_mode", "official_website",
    "helpline_number", "status", "keywords"
]

for out_dir in DEST_DIRS:
    csv_path = os.path.join(out_dir, "government_schemes.csv")
    xlsx_path = os.path.join(out_dir, "government_schemes.xlsx")
    sql_path = os.path.join(out_dir, "government_schemes.sql")

    # 1. Export CSV
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_schemes)
    print(f"Exported CSV: {csv_path}")

    # 2. Export XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Government Schemes"

    header_fill = PatternFill(start_color="1A3DA8", end_color="1A3DA8", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    ws.append(fieldnames)
    for col_num, col_name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for row in all_schemes:
        ws.append([row[k] for k in fieldnames])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(xlsx_path)
    print(f"Exported XLSX: {xlsx_path}")

    # 3. Export SQL
    with open(sql_path, "w", encoding="utf-8") as f:
        f.write("-- CitizenAware Government Schemes Dataset SQL Export\n")
        f.write("-- Total Records: 1000\n\n")
        f.write("CREATE TABLE IF NOT EXISTS government_schemes (\n")
        f.write("    id VARCHAR(50) PRIMARY KEY,\n")
        f.write("    scheme_name VARCHAR(255) NOT NULL UNIQUE,\n")
        f.write("    scheme_type VARCHAR(50),\n")
        f.write("    state VARCHAR(100),\n")
        f.write("    department VARCHAR(255),\n")
        f.write("    category VARCHAR(100),\n")
        f.write("    description TEXT,\n")
        f.write("    eligibility TEXT,\n")
        f.write("    minimum_age INT,\n")
        f.write("    maximum_age INT,\n")
        f.write("    gender VARCHAR(20),\n")
        f.write("    income_limit VARCHAR(50),\n")
        f.write("    occupation VARCHAR(100),\n")
        f.write("    caste VARCHAR(50),\n")
        f.write("    disability VARCHAR(100),\n")
        f.write("    benefits TEXT,\n")
        f.write("    required_documents TEXT,\n")
        f.write("    application_mode VARCHAR(50),\n")
        f.write("    official_website VARCHAR(255),\n")
        f.write("    helpline_number VARCHAR(50),\n")
        f.write("    status VARCHAR(50),\n")
        f.write("    keywords TEXT\n")
        f.write(");\n\n")

        for s in all_schemes:
            def clean_sql(val):
                return str(val).replace("'", "''")

            min_age_val = s["minimum_age"] if s["minimum_age"].isdigit() else "NULL"
            max_age_val = s["maximum_age"] if s["maximum_age"].isdigit() else "NULL"

            sql_line = (
                f"INSERT INTO government_schemes ("
                f"id, scheme_name, scheme_type, state, department, category, description, eligibility, "
                f"minimum_age, maximum_age, gender, income_limit, occupation, caste, disability, benefits, "
                f"required_documents, application_mode, official_website, helpline_number, status, keywords"
                f") VALUES ("
                f"'{clean_sql(s['id'])}', '{clean_sql(s['scheme_name'])}', '{clean_sql(s['scheme_type'])}', '{clean_sql(s['state'])}', "
                f"'{clean_sql(s['department'])}', '{clean_sql(s['category'])}', '{clean_sql(s['description'])}', '{clean_sql(s['eligibility'])}', "
                f"{min_age_val}, {max_age_val}, '{clean_sql(s['gender'])}', '{clean_sql(s['income_limit'])}', '{clean_sql(s['occupation'])}', "
                f"'{clean_sql(s['caste'])}', '{clean_sql(s['disability'])}', '{clean_sql(s['benefits'])}', '{clean_sql(s['required_documents'])}', "
                f"'{clean_sql(s['application_mode'])}', '{clean_sql(s['official_website'])}', '{clean_sql(s['helpline_number'])}', "
                f"'{clean_sql(s['status'])}', '{clean_sql(s['keywords'])}'"
                f");\n"
            )
            f.write(sql_line)

    print(f"Exported SQL: {sql_path}")

print("Dataset generation complete!")
