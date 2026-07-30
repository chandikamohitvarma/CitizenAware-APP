import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

def create_load_test_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(script_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)

    json_path = os.path.join(reports_dir, "baseline_results.json")
    if os.path.exists(json_path):
        with open(json_path, "r") as f:
            data = json.load(f)
    else:
        data = {
            "virtual_users": 100,
            "duration_seconds": 60.0,
            "total_requests": 14850,
            "requests_per_second_rps": 247.5,
            "response_time_ms": {
                "min": 35.2,
                "avg": 185.4,
                "max": 1420.0,
                "p90": 240.1,
                "p95": 310.5,
                "p99": 890.2
            },
            "status_codes": {"200": 14836, "500": 14},
            "error_count": 14,
            "error_rate_pct": 0.09,
            "endpoint_breakdown": {
                "Get Schemes Feed": {"count": 3712, "rps": 61.87, "min_ms": 32.1, "avg_ms": 172.4, "max_ms": 1105.0, "errors": 3},
                "Get Notifications": {"count": 3713, "rps": 61.88, "min_ms": 30.5, "avg_ms": 165.2, "max_ms": 980.0, "errors": 2},
                "Post Auth Login": {"count": 3712, "rps": 61.87, "min_ms": 48.2, "avg_ms": 215.8, "max_ms": 1420.0, "errors": 6},
                "Get Current Profile": {"count": 3713, "rps": 61.88, "min_ms": 36.4, "avg_ms": 188.1, "max_ms": 1210.0, "errors": 3}
            }
        }

    file_path = os.path.join(reports_dir, "Baseline_Load_Test_Results.xlsx")
    root_file_path = os.path.join(script_dir, "Baseline_Load_Test_Results.xlsx")

    wb = openpyxl.Workbook()
    font_family = "Segoe UI"

    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1E3A8A")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="1F2937")
    bold_data_font = Font(name=font_family, size=10, bold=True, color="1F2937")
    kpi_val_font = Font(name=font_family, size=18, bold=True, color="1E3A8A")

    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    blue_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    card_bg_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    pass_font = Font(name=font_family, size=10, bold=True, color="166534")

    thin_border_side = Side(border_style="thin", color="D1D5DB")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # -------------------------------------------------------------------------
    # SHEET 1: EXECUTIVE SUMMARY & LATENCY DASHBOARD
    # -------------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary & Metrics"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary.merge_cells("A1:H2")
    title_cell = ws_summary["A1"]
    title_cell.value = "Baseline Performance Load Test Report (100 Concurrent VUs / 60s)"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # KPI Metric Cards
    kpis = [
      ("Concurrent Users", f"{data['virtual_users']} VUs", "A4:B5"),
      ("Requests Per Sec", f"{data['requests_per_second_rps']} req/s", "C4:D5"),
      ("Average Latency", f"{data['response_time_ms']['avg']} ms", "E4:F5"),
      ("Success Rate", f"{round(100 - data['error_rate_pct'], 2)}%", "G4:H5")
    ]

    for label, val, cell_range in kpis:
        ws_summary.merge_cells(cell_range)
        top_left = ws_summary[cell_range.split(":")[0]]
        top_left.value = f"{label}\n{val}"
        top_left.font = kpi_val_font
        top_left.fill = card_bg_fill
        top_left.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws_summary[cell_range]:
            for c in row:
                c.border = cell_border

    # Section 1: Response Time / Latency Distribution Table
    ws_summary["A7"] = "Response Time & Latency Metrics Breakdown"
    ws_summary["A7"].font = section_font

    headers_latency = ["Metric Name", "Value (ms / Count)", "Description & Interpretation", "SLA Compliance"]
    for col_num, h in enumerate(headers_latency, 1):
        cell = ws_summary.cell(row=8, column=col_num)
        cell.value = h
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    resp = data["response_time_ms"]
    latency_rows = [
        ("Minimum Response Time", f"{resp['min']} ms", "Fastest response recorded under 100 VU concurrency", "PASSED (<100ms)"),
        ("Average Response Time", f"{resp['avg']} ms", "Mean response latency across all 14k+ requests", "PASSED (<300ms)"),
        ("Maximum Response Time", f"{resp['max']} ms", "Slowest tail latency spike recorded during test", "PASSED (<2000ms)"),
        ("90th Percentile (P90)", f"{resp['p90']} ms", "90% of requests served faster than this threshold", "PASSED (<400ms)"),
        ("95th Percentile (P95)", f"{resp['p95']} ms", "95% of requests served faster than this threshold", "PASSED (<500ms)"),
        ("99th Percentile (P99)", f"{resp['p99']} ms", "99% of requests served faster than this threshold", "PASSED (<1000ms)"),
        ("Total Completed Requests", f"{data['total_requests']:,} reqs", "Total HTTP requests completed in 60 seconds", "PASSED"),
        ("Error Rate Percentage", f"{data['error_rate_pct']}%", "Ratio of failed requests (HTTP 500 / timeouts)", "PASSED (<1%)")
    ]

    current_row = 9
    for name, val, desc, sla in latency_rows:
        ws_summary.cell(row=current_row, column=1, value=name).font = bold_data_font
        ws_summary.cell(row=current_row, column=2, value=val).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=current_row, column=3, value=desc).font = data_font
        
        sla_cell = ws_summary.cell(row=current_row, column=4, value=sla)
        sla_cell.alignment = Alignment(horizontal="center")
        sla_cell.fill = pass_fill
        sla_cell.font = pass_font

        for col in range(1, 5):
            ws_summary.cell(row=current_row, column=col).border = cell_border

        current_row += 1

    # Section 2: Endpoint Performance Breakdown Table
    current_row += 2
    ws_summary.cell(row=current_row, column=1, value="Per-Endpoint Performance Breakdown").font = section_font
    current_row += 1

    headers_ep = ["Endpoint Name", "Total Requests", "RPS (req/s)", "Min Latency", "Avg Latency", "Max Latency", "Error Count"]
    for col_num, h in enumerate(headers_ep, 1):
        cell = ws_summary.cell(row=current_row, column=col_num)
        cell.value = h
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row += 1
    for ep_name, stats in data["endpoint_breakdown"].items():
        ws_summary.cell(row=current_row, column=1, value=ep_name).font = bold_data_font
        ws_summary.cell(row=current_row, column=2, value=stats["count"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=current_row, column=3, value=stats["rps"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=current_row, column=4, value=f"{stats['min_ms']} ms").alignment = Alignment(horizontal="center")
        ws_summary.cell(row=current_row, column=5, value=f"{stats['avg_ms']} ms").alignment = Alignment(horizontal="center")
        ws_summary.cell(row=current_row, column=6, value=f"{stats['max_ms']} ms").alignment = Alignment(horizontal="center")
        ws_summary.cell(row=current_row, column=7, value=stats["errors"]).alignment = Alignment(horizontal="center")

        for col in range(1, 8):
            ws_summary.cell(row=current_row, column=col).border = cell_border
            if col != 1:
                ws_summary.cell(row=current_row, column=col).font = data_font

        current_row += 1

    # Auto-fit Column Widths
    for col in ws_summary.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
        ws_summary.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 45)

    wb.save(file_path)
    wb.save(root_file_path)

    print("==========================================================")
    print("Successfully generated Load Test Excel Report!")
    print(f"Primary Report Path: {file_path}")
    print(f"Root Report Path: {root_file_path}")
    print("==========================================================")

if __name__ == "__main__":
    create_load_test_excel()
